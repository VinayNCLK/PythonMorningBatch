import openpyxl


def getnoofrows(path,sheetname):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetname]
    return sheet.max_row


def getnoofcolumns(path,sheetname):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetname]
    return sheet.max_column


def getcellvalue(path,sheetname,rowno,colmunno):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetname]
    cell = sheet.cell(row=rowno,column=colmunno)
    return cell.value


def writecellvalue(path,sheetname,rowno,colmunno,value):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetname]
    cell = sheet.cell(row=rowno, column=colmunno)
    cell.value = value
    wb.save(path)
